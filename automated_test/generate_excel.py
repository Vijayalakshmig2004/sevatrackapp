import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

def main():
    report_json_path = os.path.join(os.path.dirname(__file__), "report.json")
    if not os.path.exists(report_json_path):
        print("report.json not found!")
        return

    with open(report_json_path, "r") as f:
        results = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAST Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Summary
    ws.append(["=== DAST SUMMARY ==="])
    ws.append(["Total Tests", "Findings", "Passed", "Failed"])
    
    total = len(results)
    failed = sum(1 for r in results if r.get("finding", False))
    passed = total - failed

    ws.append([total, failed, passed, failed])

    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill

    ws.append([]) # Empty row

    # Details
    ws.append(["=== TEST DETAILS ==="])
    headers = ["Endpoint", "Method", "Category", "Status", "Expected", "Finding", "Severity", "Time (ms)", "Note"]
    ws.append(headers)

    for cell in ws[6]:
        cell.font = header_font
        cell.fill = header_fill

    for r in results:
        row = [
            r.get("endpoint"),
            r.get("method"),
            r.get("test_category"),
            r.get("status"),
            r.get("expected_status"),
            "FAIL" if r.get("finding") else "PASS",
            r.get("severity"),
            r.get("response_time_ms"),
            r.get("note")
        ]
        ws.append(row)
        
        # Color coding
        finding_cell = ws.cell(row=ws.max_row, column=6)
        if r.get("finding"):
            finding_cell.font = Font(color="FF0000", bold=True)
        else:
            finding_cell.font = Font(color="00B050", bold=True)

    # Adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    report_excel_path = os.path.join(os.path.dirname(__file__), "DAST_Report_Updated.xlsx")
    wb.save(report_excel_path)
    print(f"Excel report generated at {report_excel_path}")

if __name__ == "__main__":
    main()
